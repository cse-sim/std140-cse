print("\nInitializing Python...\n")
import openpyxl as xl
import pandas as pd
from datetime import datetime
import mako.template as mk
import os
import pytz
import os, glob


def call_csv(path):
    data = pd.read_csv(path)
    return pd.DataFrame(data)


# delete output file
def delete_output_file(path):
    for file in glob.glob(f"{path}*"):
        os.remove(file)


# prints date and time with time zone on each plot
def find_todays_date():
    utc_timezone = pytz.timezone("America/Denver")
    current_date_time = datetime.now(utc_timezone)
    return current_date_time.strftime("%Y%m%d")


output_file_root = "CSE-ET100series"
tests = "etna"
template_file = "ET100series-Output-GMT+1 (071023a).xlsx"
current_directory = os.path.dirname(os.path.dirname(os.getcwd()))
# current_directory = os.getcwd()
delete_output_file(f"{current_directory}/reports/etna/{output_file_root}")
output_file = f"{output_file_root}-{find_todays_date()}"
template = xl.load_workbook(filename=f"{current_directory}/docs/etna/{template_file}")
cases = template.sheetnames
case_results_path = "output/etna"
cases_cell_A = ["ET100A1", "ET100A3", "ET110A1", "ET110A2"]
cases_insulated_windows = ["ET110A1", "ET110A2", "ET110B1", "ET110B2"]

experiment_dates = {
    "cases_insulated_windows": {
        "cases_cell_A": {
            "start_date": datetime(2000, 1, 26, 12),
            "end_date": datetime(2000, 2, 11, 9),
        },
        "cases_cell_B": {
            "start_date": datetime(2000, 1, 26, 12),
            "end_date": datetime(2000, 2, 11, 9),
        },
    },
    "cases_uninsulated_windows": {
        "cases_cell_A": {
            "start_date": datetime(2000, 9, 8, 16),
            "end_date": datetime(2000, 9, 18, 14),
        },
        "cases_cell_B": {
            "start_date": datetime(2000, 9, 8, 16),
            "end_date": datetime(2000, 9, 18, 15),
        },
    },
}

output_columns = {
    "cases_cell_A": {
        "CeilingPath1 Exterior Surface Temperature - Simulation [C]": 2,
        "FloorPath1 Exterior Surface Temperature - Simulation [C]": 3,
        "NorthWall Exterior Surface Temperature - Simulation [C]": 4,
        "EastWall Exterior Surface Temperature - Simulation [C]": 5,
        "SouthWall Exterior Surface Temperature - Simulation [C]": 6,
        "WestWall Exterior Surface Temperature - Simulation [C]": 7,
        "Cell Temperature - Simulation [C]": 8,
        "Fan-Simulation [Wh]": 9,
        "Blower Fan Flow Rate-Simulation [m3/h]": 10,
        "Heater-Simulation [Wh]": 11,
        "CeilingPath1 Convective Heat Flux [Wh/m2]": 12,
        "CeilingPath1 Radiative Heat Flux [Wh/m2]": 13,
        "CeilingPath1 Total Heat Flux [Wh/m2]": 14,
        "CeilingPath2 Convective Heat Flux [Wh/m2]": 15,
        "CeilingPath2 Radiative Heat Flux [Wh/m2]": 16,
        "CeilingPath2 Total Heat Flux [Wh/m2]": 17,
        "FloorPath1 Convective Heat Flux [Wh/m2]": 18,
        "FloorPath1 Radiative Heat Flux [Wh/m2]": 19,
        "FloorPath1 Total Heat Flux [Wh/m2]": 20,
        "FloorPath2 Convective Heat Flux [Wh/m2]": 21,
        "FloorPath2 Radiative Heat Flux [Wh/m2]": 22,
        "FloorPath2 Total Heat Flux [Wh/m2]": 23,
        "FloorPath3 Convective Heat Flux [Wh/m2]": 24,
        "FloorPath3 Radiative Heat Flux [Wh/m2]": 25,
        "FloorPath3 Total Heat Flux [Wh/m2]": 26,
        "NorthWall Convective Heat Flux [Wh/m2]": 27,
        "NorthWall Radiative Heat Flux [Wh/m2]": 28,
        "NorthWall Total Heat Flux [Wh/m2]": 29,
        "NorthWallDoor Convective Heat Flux [Wh/m2]": 30,
        "NorthWallDoor Radiative Heat Flux [Wh/m2]": 31,
        "NorthWallDoor Total Heat Flux [Wh/m2]": 32,
        "EastWall Convective Heat Flux [Wh/m2]": 33,
        "EastWall Radiative Heat Flux [Wh/m2]": 34,
        "EastWall Total Heat Flux [Wh/m2]": 35,
        "WindowsPath1East Convective Heat Flux [Wh/m2]": 36,
        "WindowsPath1East Radiative Heat Flux [Wh/m2]": 37,
        "WindowsPath1East Total Heat Flux [Wh/m2]": 38,
        "WindowsPath2East Convective Heat Flux [Wh/m2]": 39,
        "WindowsPath2East Radiative Heat Flux [Wh/m2]": 40,
        "WindowsPath2East Total Heat Flux [Wh/m2]": 41,
        "WindowsPath3East Convective Heat Flux [Wh/m2]": 42,
        "WindowsPath3East Radiative Heat Flux [Wh/m2]": 43,
        "WindowsPath3East Total Heat Flux [Wh/m2]": 44,
        "SouthWall Convective Heat Flux [Wh/m2]": 45,
        "SouthWall Radiative Heat Flux [Wh/m2]": 46,
        "SouthWall Total Heat Flux [Wh/m2]": 47,
        "WindowsPath1South Convective Heat Flux [Wh/m2]": 48,
        "WindowsPath1South Radiative Heat Flux [Wh/m2]": 49,
        "WindowsPath1South Total Heat Flux [Wh/m2]": 50,
        "WindowsPath2South Convective Heat Flux [Wh/m2]": 51,
        "WindowsPath2South Radiative Heat Flux [Wh/m2]": 52,
        "WindowsPath2South Total Heat Flux [Wh/m2]": 53,
        "WindowsPath3South Convective Heat Flux [Wh/m2]": 54,
        "WindowsPath3South Radiative Heat Flux [Wh/m2]": 55,
        "WindowsPath3South Total Heat Flux [Wh/m2]": 56,
        "WestWall Convective Heat Flux [Wh/m2]": 57,
        "WestWall Radiative Heat Flux [Wh/m2]": 58,
        "WestWall Total Heat Flux [Wh/m2]": 59,
    },
    "cases_cell_B": {
        "CeilingPath1 Exterior Surface Temperature - Simulation [C]": 2,
        "FloorPath1 Exterior Surface Temperature - Simulation [C]": 3,
        "NorthWall Exterior Surface Temperature - Simulation [C]": 4,
        "EastWall Exterior Surface Temperature - Simulation [C]": 5,
        "SouthWall Exterior Surface Temperature - Simulation [C]": 6,
        "WestWall Exterior Surface Temperature - Simulation [C]": 7,
        "Cell Temperature - Simulation [C]": 8,
        "Fan-Simulation [Wh]": 9,
        "Blower Fan Flow Rate-Simulation [m3/h]": 10,
        "Heater-Simulation [Wh]": 11,
        "CeilingPath1 Convective Heat Flux [Wh/m2]": 12,
        "CeilingPath1 Radiative Heat Flux [Wh/m2]": 13,
        "CeilingPath1 Total Heat Flux [Wh/m2]": 14,
        "CeilingPath2 Convective Heat Flux [Wh/m2]": 15,
        "CeilingPath2 Radiative Heat Flux [Wh/m2]": 16,
        "CeilingPath2 Total Heat Flux [Wh/m2]": 17,
        "FloorPath1 Convective Heat Flux [Wh/m2]": 18,
        "FloorPath1 Radiative Heat Flux [Wh/m2]": 19,
        "FloorPath1 Total Heat Flux [Wh/m2]": 20,
        "FloorPath2 Convective Heat Flux [Wh/m2]": 21,
        "FloorPath2 Radiative Heat Flux [Wh/m2]": 22,
        "FloorPath2 Total Heat Flux [Wh/m2]": 23,
        "FloorPath3 Convective Heat Flux [Wh/m2]": 24,
        "FloorPath3 Radiative Heat Flux [Wh/m2]": 25,
        "FloorPath3 Total Heat Flux [Wh/m2]": 26,
        "NorthWall Convective Heat Flux [Wh/m2]": 27,
        "NorthWall Radiative Heat Flux [Wh/m2]": 28,
        "NorthWall Total Heat Flux [Wh/m2]": 29,
        "NorthWallDoor Convective Heat Flux [Wh/m2]": 30,
        "NorthWallDoor Radiative Heat Flux [Wh/m2]": 31,
        "NorthWallDoor Total Heat Flux [Wh/m2]": 32,
        "EastWall Convective Heat Flux [Wh/m2]": 33,
        "EastWall Radiative Heat Flux [Wh/m2]": 34,
        "EastWall Total Heat Flux [Wh/m2]": 35,
        "SouthWall Convective Heat Flux [Wh/m2]": 36,
        "SouthWall Radiative Heat Flux [Wh/m2]": 37,
        "SouthWall Total Heat Flux [Wh/m2]": 38,
        "WindowsPath1South Convective Heat Flux [Wh/m2]": 39,
        "WindowsPath1South Radiative Heat Flux [Wh/m2]": 40,
        "WindowsPath1South Total Heat Flux [Wh/m2]": 41,
        "WindowsPath2South Convective Heat Flux [Wh/m2]": 42,
        "WindowsPath2South Radiative Heat Flux [Wh/m2]": 43,
        "WindowsPath2South Total Heat Flux [Wh/m2]": 44,
        "WindowsPath3South Convective Heat Flux [Wh/m2]": 45,
        "WindowsPath3South Radiative Heat Flux [Wh/m2]": 46,
        "WindowsPath3South Total Heat Flux [Wh/m2]": 47,
        "WestWall Convective Heat Flux [Wh/m2]": 48,
        "WestWall Radiative Heat Flux [Wh/m2]": 49,
        "WestWall Total Heat Flux [Wh/m2]": 50,
        "WindowsPath1West Convective Heat Flux [Wh/m2]": 51,
        "WindowsPath1West Radiative Heat Flux [Wh/m2]": 52,
        "WindowsPath1West Total Heat Flux [Wh/m2]": 53,
        "WindowsPath2West Convective Heat Flux [Wh/m2]": 54,
        "WindowsPath2West Radiative Heat Flux [Wh/m2]": 55,
        "WindowsPath2West Total Heat Flux [Wh/m2]": 56,
        "WindowsPath3West Convective Heat Flux [Wh/m2]": 57,
        "WindowsPath3West Radiative Heat Flux [Wh/m2]": 58,
        "WindowsPath3West Total Heat Flux [Wh/m2]": 59,
    },
}

output_cases_interior_surface_temperatures = {
    "cases_cell_A": {
        "CeilingPath1 Interior Surface Temperature - Simulation [C]": 60,
        "CeilingPath2 Interior Surface Temperature - Simulation [C]": 61,
        "FloorPath1 Interior Surface Temperature - Simulation [C]": 62,
        "FloorPath2 Interior Surface Temperature - Simulation [C]": 63,
        "FloorPath3 Interior Surface Temperature - Simulation [C]": 64,
        "NorthWall Interior Surface Temperature - Simulation [C]": 65,
        "NorthWallDoor Interior Surface Temperature - Simulation [C]": 66,
        "EastWall Interior Surface Temperature - Simulation [C]": 67,
        "WindowsPath1East Interior Surface Temperature - Simulation [C]": 68,
        "WindowsPath2East Interior Surface Temperature - Simulation [C]": 69,
        "WindowsPath3East Interior Surface Temperature - Simulation [C]": 70,
        "SouthWall Interior Surface Temperature - Simulation [C]": 71,
        "WindowsPath1South Interior Surface Temperature - Simulation [C]": 72,
        "WindowsPath2South Interior Surface Temperature - Simulation [C]": 73,
        "WindowsPath3South Interior Surface Temperature - Simulation [C]": 74,
        "WestWall Interior Surface Temperature - Simulation [C]": 75,
    },
    "cases_cell_B": {
        "CeilingPath1 Interior Surface Temperature - Simulation [C]": 60,
        "CeilingPath2 Interior Surface Temperature - Simulation [C]": 61,
        "FloorPath1 Interior Surface Temperature - Simulation [C]": 62,
        "FloorPath2 Interior Surface Temperature - Simulation [C]": 63,
        "FloorPath3 Interior Surface Temperature - Simulation [C]": 64,
        "NorthWall Interior Surface Temperature - Simulation [C]": 65,
        "NorthWallDoor Interior Surface Temperature - Simulation [C]": 66,
        "EastWall Interior Surface Temperature - Simulation [C]": 67,
        "SouthWall Interior Surface Temperature - Simulation [C]": 68,
        "WindowsPath1South Interior Surface Temperature - Simulation [C]": 69,
        "WindowsPath2South Interior Surface Temperature - Simulation [C]": 70,
        "WindowsPath3South Interior Surface Temperature - Simulation [C]": 71,
        "WestWall Interior Surface Temperature - Simulation [C]": 72,
        "WindowsPath1West Interior Surface Temperature - Simulation [C]": 73,
        "WindowsPath2West Interior Surface Temperature - Simulation [C]": 74,
        "WindowsPath3West Interior Surface Temperature - Simulation [C]": 75,
    },
}


def find_case_specifics(case):
    global experiment_dates
    global cases_insulated_windows
    global cases_cell_A

    if case in cases_cell_A:
        cell = "cases_cell_A"
    else:
        cell = "cases_cell_B"
    if case in cases_insulated_windows:
        start_date, end_date = (
            experiment_dates["cases_insulated_windows"][cell]["start_date"],
            experiment_dates["cases_insulated_windows"][cell]["end_date"],
        )
    else:
        start_date, end_date = (
            experiment_dates["cases_uninsulated_windows"][cell]["start_date"],
            experiment_dates["cases_uninsulated_windows"][cell]["end_date"],
        )
    return cell, start_date, end_date


for case in cases:
    cell, start_date, end_date = find_case_specifics(case)
    case_template = template[case]
    case_results = call_csv(
        f"{current_directory}/{case_results_path}/{case}/OUTPUT.CSV"
    )
    case_results["Datetime"] = [
        datetime(
            2000,
            case_results.loc[index, "Month"],
            case_results.loc[index, "Day"],
            case_results.loc[index, "Hour"] - 1,
        )
        for index in range(len(case_results))
    ]
    case_results = case_results[
        (case_results["Datetime"] >= start_date)
        & (case_results["Datetime"] <= end_date)
    ].reset_index()

    for index in range(len(case_results)):
        if case in ["ET100A3", "ET100B3"]:
            output_columns_final = output_columns[cell].copy()
            output_columns_final.update(
                output_cases_interior_surface_temperatures[cell]
            )
        else:
            output_columns_final = output_columns[cell].copy()
        for output_column, template_column in output_columns_final.items():
            case_template.cell(column=template_column, row=index + 6).value = (
                case_results.loc[index, output_column]
            )

print("Done processing cases.")
print("Writing results to XLSX ...")

template.save(filename=f"{current_directory}/reports/{tests}/{output_file}.xlsx")
