print("\nInitializing Python...\n")
import openpyxl as xl
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from datetime import timedelta, datetime, date
import subprocess
import re
import mako.template as mk
import os

#%matplotlib inline

tests = 'weather-drivers'
# os.chdir('scripts/%s' % (tests)) # use only for Hydrogen testing in Atom
wb = xl.load_workbook(filename='../../reports/' + tests + '/WeatherDriversResultsSubmittal-Template.xlsx')
Sheet1 = wb['Sheet1']

# Top level info
info = str(subprocess.check_output("..\\..\\CSE.exe", shell=True))
match = re.compile('.*CSE\s+([^\s]*)\s+.*',re.S).match(info)
if match:
  version = match.groups()[0]
else:
  print("ERROR: Unknown version!")
  version = "????"

Sheet1.cell(column = 2, row = 3).value = "CSE v{}".format(version) # Program Name and Version:
Sheet1.cell(column = 2, row = 4).value = "2020-03-12" # Program Version Release Date:
Sheet1.cell(column = 2, row = 5).value = "CSE" # Program Name for Tables and Charts (short):
Sheet1.cell(column = 2, row = 6).value = "Big Ladder Software" # Modeler Organization:
Sheet1.cell(column = 2, row = 7).value = "BLS" # Modeler Organization for Tables and Charts (short):
Sheet1.cell(column = 2, row = 8).value = '{}'.format(datetime.now()) # Results Submission Date:

# Annual Outputs
annual_row_beg = 12 # Average Dry Bulb Temperature (C)
annual_row_end = 46 # Total Diffuse Radiation on W Azimuth and 30° from H Slope (Wh/m2)
annual_init_col = 3 # 'WD100'
annual_end_col = 8 # 'WD600'

col_surfaces = [
  "Total Horizontal Radiation (Wh/m2)",
  "Beam Horizontal Radiation (Wh/m2)",
  "Diffuse Horizontal Radiation (Wh/m2)",
  "Total Radiation on S Azimuth and 90° Slope (Wh/m2)",
  "Total Beam Radiation on S Azimuth and 90° Slope (Wh/m2)",
  "Total Diffuse Radiation on S Azimuth and 90° Slope (Wh/m2)",
  "Total Radiation on E Azimuth and 90° Slope (Wh/m2)",
  "Total Beam Radiation on E Azimuth and 90° Slope (Wh/m2)",
  "Total Diffuse Radiation on E Azimuth and 90° Slope (Wh/m2)",
  "Total Radiation on N Azimuth and 90° Slope (Wh/m2)",
  "Total Beam Radiation on N Azimuth and 90° Slope (Wh/m2)",
  "Total Diffuse Radiation on N Azimuth and 90° Slope (Wh/m2)",
  "Total Radiation on W Azimuth and 90° Slope (Wh/m2)",
  "Total Beam Radiation on W Azimuth and 90° Slope (Wh/m2)",
  "Total Diffuse Radiation on W Azimuth and 90° Slope (Wh/m2)",
  "Total Radiation on 45° E of S Azimuth and 90° Slope (Wh/m2)",
  "Total Beam Radiation on 45° E of S Azimuth and 90° Slope (Wh/m2)",
  "Total Diffuse Radiation on 45° E of S Azimuth and 90° Slope (Wh/m2)",
  "Total Radiation on 45° W of S Azimuth and 90° Slope (Wh/m2)",
  "Total Beam Radiation on 45° W of S Azimuth and 90° Slope (Wh/m2)",
  "Total Diffuse Radiation on 45° W of S Azimuth and 90° Slope (Wh/m2)",
  "Total Radiation on E Azimuth and 30° from H Slope (Wh/m2)",
  "Total Beam Radiation on E Azimuth and 30° from H Slope (Wh/m2)",
  "Total Diffuse Radiation on E Azimuth and 30° from H Slope (Wh/m2)",
  "Total Radiation on S Azimuth and 30° from H Slope (Wh/m2)",
  "Total Beam Radiation on S Azimuth and 30° from H Slope (Wh/m2)",
  "Total Diffuse Radiation on S Azimuth and 30° from H Slope (Wh/m2)",
  "Total Radiation on W Azimuth and 30° from H Slope (Wh/m2)",
  "Total Beam Radiation on W Azimuth and 30° from H Slope (Wh/m2)",
  "Total Diffuse Radiation on W Azimuth and 30° from H Slope (Wh/m2)"
]

col_hourly = [
  "Dry Bulb Temperature (C)",
  "Relative Humidity (%)",
  "Dewpoint Temperature (C)",
  "Humidity Ratio (kg moisture/kg dry air)",
  "Wet Bulb Temperature (C)",
  "Windspeed (m/s)",
  "Wind Direction (degrees from North)",
  "Station Pressure (mbar)",
  "Total Cloud Cover (tenths of sky)",
  "Opaque Cloud Cover (tenths of sky)",
  "Sky Temperature (C)"
]

col_subhourly = [
  "Dry Bulb Temperature (C)",
  "Relative Humidity (%)"
]

# map case_name to requested days to generate outputs for
report_days = {
  "WD100": [[5, 4], [7, 14], [9, 6]], #["MAY 4", "JUL 14", "SEP 6"]
  "WD200": [[5, 24], [8, 26]], # ["MAY 24", "AUG 26"]
  "WD300": [[2, 7], [8, 13]], # ["FEB 7", "AUG 13"]
  "WD400": [[1, 24], [7, 1]], # ["JAN 24", "JUL 1"]
  "WD500": [[3, 1], [9, 14]], # ["MAR 1", "SEP 14"]
  "WD600": [[5, 4], [7, 14], [9, 6]] # ["MAY 4", "JUL 14", "SEP 6"]
}

print("Adding empty rows to results spreadsheet...")
# fixes missing two rows in some hourly outputs
for row in [939,940]:
  Sheet1.insert_rows(row)

# add empty rows for subhourly outputs
timestep = 0.1 # hourly length of subhourly timesteps used by tool
subhourly_row_beg = 1124 # First block of subhourly outputs for dry bulb temperature
for output in (col_subhourly + col_surfaces):
  for row in range(subhourly_row_beg, subhourly_row_beg - 2 + len(np.arange(timestep, 24 + timestep, timestep))): # already two blank rows for each subhourly output
    Sheet1.insert_rows(row)
  subhourly_row_beg += len(np.arange(timestep, 24 + timestep, timestep)) + 2

print("Adding time of day for hourly outputs ...")

hourly_row_beg = 53 # First block of hourly outputs for dry bulb temperature
for output in (col_hourly + col_surfaces):
  for hour in range(1, 25):
    Sheet1.cell(column = 1, row = hourly_row_beg + hour - 1).value = hour
  hourly_row_beg += 26

print("Adding time of day for subhourly outputs ...")

subhourly_row_beg = 1124 # First block of subhourly outputs for dry bulb temperature
for output in (col_subhourly + col_surfaces):
  i = 0
  for subhour in np.arange(timestep, 24 + timestep, timestep):
    Sheet1.cell(column = 1, row = subhourly_row_beg + i).value = subhour
    i += 1
  subhourly_row_beg += len(np.arange(timestep, 24 + timestep, timestep)) + 2

print("Processing case: ")

for col in range(annual_init_col, annual_end_col + 1):
  row  = annual_row_beg
  case = str(Sheet1.cell(column = col, row = row - 1).value)
  print("  " + case)
  df_hourly = pd.read_csv('../../output/' + tests + '/' + case + '/HOURLY.csv')
  df_subhourly = pd.read_csv('../../output/' + tests + '/' + case + '/SUBHOURLY.csv')

  # Annual average outdoor air data
  Sheet1.cell(column = col, row = row).value = df_subhourly['Dry Bulb Temperature (C)'].mean()
  Sheet1.cell(column = col, row = row + 2).value = df_hourly['Dewpoint Temperature (C)'].mean()
  Sheet1.cell(column = col, row = row + 3).value = df_hourly['Humidity Ratio (kg moisture/kg dry air)'].mean()
  Sheet1.cell(column = col, row = row + 4).value = df_hourly['Wet Bulb Temperature (C)'].mean()

  # Annual sums for incident solar radiation on surfaces
  i = 0
  while i < len(col_surfaces):
    Sheet1.cell(column = col, row = row + 5 + i).value = df_subhourly[col_surfaces[i]].sum()
    i += 1

  # Hourly Outputs
  hourly_init_col = 2 # 'WD100' for May 4
  for key in report_days:
    if case == key:
      for day in report_days[key]:
        hourly_row_beg = 53 # First block of hourly outputs for dry bulb temperature
        df_day_hr = df_hourly[(df_hourly['Month'] == day[0]) & (df_hourly['Day'] == day[1])]
        for i in range(1, len(col_hourly) + 1): # iterate over hourly outputs for outdoor air conditions
          if col_hourly[i - 1] in df_day_hr.columns: # skip outputs not created by tool
            for hour in range(1, 25):
              df_hour_hr = df_day_hr[df_day_hr['Hour'] == hour]
              row_i = hourly_row_beg + hour - 1
              Sheet1.cell(column = hourly_init_col, row = row_i).value = df_hour_hr[col_hourly[i - 1]].mean()
          hourly_row_beg += 26
        df_day_shr = df_subhourly[(df_subhourly['Month'] == day[0]) & (df_subhourly['Day'] == day[1])]
        for i in range(1, len(col_surfaces) + 1): # iterate over subhourly outputs for surface radiation
          for hour in range(1, 25):
            df_hour_shr = df_day_shr[df_day_shr['Hour'] == hour]
            row_i = hourly_row_beg + hour - 1
            Sheet1.cell(column = hourly_init_col, row = row_i).value = df_hour_shr[col_surfaces[i - 1]].sum()
            if i < 4:
              Sheet1.cell(column = hourly_init_col, row = row_i + 8871 - 339).value = df_hour_shr[col_surfaces[i - 1]].sum() # copy sum to final rows asking for subhourly integration of horizontal roof radiation
          hourly_row_beg += 26
        hourly_init_col += 1
    else:
      hourly_init_col = hourly_init_col + len(report_days[key])

  # Subhourly Outputs - Instantaneous timestep values
  subhourly_init_col = 2 # 'WD100' for May 4
  for key in report_days:
    if case == key:
      for day in report_days[key]:
        subhourly_row_beg = 1124 # First block of subhourly outputs for dry bulb temperature
        df_day_shr = df_subhourly[(df_subhourly['Month'] == day[0]) & (df_subhourly['Day'] == day[1])]
        for i in range(1, len(col_subhourly) + 1): # iterate over hourly outputs for outdoor air conditions
          if col_subhourly[i - 1] in df_day_shr.columns: # skip outputs not created by tool
            for hour in range(1, 25):
              df_hour_shr = df_day_shr[df_day_shr['Hour'] == hour]
              for subhour in range(1, 11): # 10 timesteps each hour
                df_subhour_shr = df_hour_shr[df_hour_shr['Subhour'] == subhour]
                row_i = subhourly_row_beg + (hour - 1)*10 + subhour - 1
                Sheet1.cell(column = subhourly_init_col, row = row_i).value = df_subhour_shr[col_subhourly[i - 1]].mean()
          subhourly_row_beg += len(np.arange(timestep, 24 + timestep, timestep)) + 2
        for i in range(1, len(col_surfaces) + 1): # iterate over subhourly outputs for surface radiation
          for hour in range(1, 25):
            df_hour_shr = df_day_shr[df_day_shr['Hour'] == hour]
            for subhour in range(1, 11): # 10 timesteps each hour
              df_subhour_shr = df_hour_shr[df_hour_shr['Subhour'] == subhour]
              row_i = subhourly_row_beg + (hour - 1)*10 + subhour - 1
              Sheet1.cell(column = subhourly_init_col, row = row_i).value = df_subhour_shr[col_surfaces[i - 1]].mean()
          subhourly_row_beg += len(np.arange(timestep, 24 + timestep, timestep)) + 2
        subhourly_init_col += 1
    else:
      subhourly_init_col = subhourly_init_col + len(report_days[key])

  # Subhourly Outputs - Hourly sums of instantaneous timestep values
  # subhourly_init_col = 2 # 'WD100' for May 4
  # for key in report_days:
  #   if case == key:
  #     for day in report_days[key]:
  #       subhourly_row_beg = 8871 # First block of subhourly outputs for total horizontal radiation after inserting empty rows
  #       df_day_shr = df_subhourly[(df_subhourly['Month'] == day[0]) & (df_subhourly['Day'] == day[1])]
  #       for i in range(1, len(col_surfaces[0..2])): # iterate over first three surface radiation subhourly outputs
  #         for hour in range(1, 25):
  #           df_hour_shr = df_day_shr[df_day_shr['Hour'] == hour]
  #           for subhour in range(1, 11): # 10 timesteps each hour
  #               df_subhour_shr = df_hour_shr[df_hour_shr['Subhour'] == subhour]
  #               row_i = hourly_row_beg + (hour - 1)*10 + subhour - 1
  #               # print(" row_i: {}".format(row_i))
  #               Sheet1.cell(column = subhourly_init_col, row = row_i).value = df_subhour_shr[col_subhourly[i - 1]].mean()
  #         subhourly_row_beg += 26
  #       hourly_init_col += 1
  #   else:
  #     hourly_init_col = hourly_init_col + len(report_days[key])



wb.save(filename='../../reports/' + tests + '/WeatherDriversResultsSubmittal.xlsx')

with open('../../reports/' + tests + '/S140outNotes-Template.txt','r') as notes_template:
  content = notes_template.read()

with open('../../reports/' + tests + '/S140outNotes.txt','w') as notes:
  notes.write(mk.Template(content).render(version=version))

print("Done")
