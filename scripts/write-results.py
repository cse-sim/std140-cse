import openpyxl as xl
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from datetime import timedelta, datetime, date

%matplotlib inline

wb = xl.load_workbook(filename='../reports/Sec5-2Aout-Empty.xlsx')
A = wb['A']
TMPBIN = wb['TMPBIN']

# Top level info
A.cell(column=3, row=61).value = "California Simulation Engine"
A.cell(column=3, row=62).value = "0.830"
A.cell(column=3, row=63).value = '{}'.format(datetime.now())

row_beg = 70 # 600
row_end = 115 # 810
row_ff_beg = 130 # 600FF
row_ff_end = 135 # 980FF
init_col = 2 # 'B'

col_loads = {
  "600": 1,
  "660": 2,
  "670": 3,
  "680": 4,
  "685": 5,
  "695": 6,
  "900": 7,
  "980": 8,
  "985": 9,
  "995": 10
}

col_temps = {
  "600FF": 21,
  "900FF": 22,
  "650FF": 23,
  "950FF": 24,
  "680FF": 25,
  "980FF": 26
}

d_month = [31,28,31,30,31,30,31,31,30,31,30,31]

print "Proccessing case: "

# Non-free-float cases
for row in range(row_beg, row_end + 1):
  case = str(A.cell(column=init_col, row=row).value)
  print "  " + case
  df = pd.read_csv('../output/' + case + '/DETAILED.csv')

  timestep = 1.0/(df['Subhour'].max() + 1)

  # Annual loads
  A.cell(column=init_col+1, row=row).value = df['Heating Load [Wh]'].sum()/1000000
  A.cell(column=init_col+2, row=row).value = df['Cooling Load [Wh]'].sum()/1000000
  idx = df['Heating Load [Wh]'].idxmax()
  A.cell(column=init_col+3, row=row).value = df.ix[idx]['Heating Load [Wh]']/1000/timestep
  A.cell(column=init_col+4, row=row).value = df.ix[idx]['Month']
  A.cell(column=init_col+5, row=row).value = df.ix[idx]['Day']
  A.cell(column=init_col+6, row=row).value = df.ix[idx]['Hour']
  idx = df['Cooling Load [Wh]'].idxmax()
  A.cell(column=init_col+7, row=row).value = df.ix[idx]['Cooling Load [Wh]']/1000/timestep
  A.cell(column=init_col+8, row=row).value = df.ix[idx]['Month']
  A.cell(column=init_col+9, row=row).value = df.ix[idx]['Day']
  A.cell(column=init_col+10, row=row).value = df.ix[idx]['Hour']

  if case == "600":
    # Solar Incidence
    A.cell(column=init_col+1, row=155).value = df['Incident Solar Roof [Wh/m2]'].sum()/1000
    A.cell(column=init_col+1, row=156).value = df['Incident Solar North Wall [Wh/m2]'].sum()/1000
    A.cell(column=init_col+1, row=157).value = df['Incident Solar East Wall [Wh/m2]'].sum()/1000
    A.cell(column=init_col+1, row=158).value = df['Incident Solar South Wall [Wh/m2]'].sum()/1000
    A.cell(column=init_col+1, row=159).value = df['Incident Solar West Wall [Wh/m2]'].sum()/1000

    # Sky Temperature
    A.cell(column=init_col+1, row=178).value = df['Sky Temp [C]'].mean()
    idx = df['Sky Temp [C]'].idxmin()
    A.cell(column=init_col+2, row=178).value = df.ix[idx]['Sky Temp [C]']
    A.cell(column=init_col+3, row=178).value = df.ix[idx]['Month']
    A.cell(column=init_col+4, row=178).value = df.ix[idx]['Day']
    A.cell(column=init_col+5, row=178).value = df.ix[idx]['Hour']
    idx = df['Sky Temp [C]'].idxmax()
    A.cell(column=init_col+6, row=178).value = df.ix[idx]['Sky Temp [C]']
    A.cell(column=init_col+7, row=178).value = df.ix[idx]['Month']
    A.cell(column=init_col+8, row=178).value = df.ix[idx]['Day']
    A.cell(column=init_col+9, row=178).value = df.ix[idx]['Hour']

  if case == "600" or case == "900":
    # Monthly loads
    if case == "600":
      col_offset = 0
    else:
      col_offset = 8

    for month in range(1,13):
        dfm = df[df['Month'] == month]
        A.cell(column=init_col+col_offset+1, row=189+month).value = dfm['Heating Load [Wh]'].sum()/1000
        A.cell(column=init_col+col_offset+2, row=189+month).value = dfm['Cooling Load [Wh]'].sum()/1000
        idx = dfm['Heating Load [Wh]'].idxmax()
        A.cell(column=init_col+col_offset+3, row=189+month).value = dfm.ix[idx]['Heating Load [Wh]']/1000/timestep
        A.cell(column=init_col+col_offset+4, row=189+month).value = dfm.ix[idx]['Day']
        A.cell(column=init_col+col_offset+5, row=189+month).value = dfm.ix[idx]['Hour']
        idx = dfm['Cooling Load [Wh]'].idxmax()
        A.cell(column=init_col+col_offset+6, row=189+month).value = dfm.ix[idx]['Cooling Load [Wh]']/1000/timestep
        A.cell(column=init_col+col_offset+7, row=189+month).value = dfm.ix[idx]['Day']
        A.cell(column=init_col+col_offset+8, row=189+month).value = dfm.ix[idx]['Hour']

  # Hourly outputs (misc.)
  if case == "600":
    dfd = df[(df['Month'] == 5) & (df['Day'] == 4)]
    for hour in range(1,25):
      dfh = dfd[dfd["Hour"] == hour]
      row_i = 230 + hour - 1

      # Solar Incidence
      A.cell(column=init_col+1, row=row_i).value = dfh['Incident Solar Roof [Wh/m2]'].sum()
      A.cell(column=init_col+2, row=row_i).value = dfh['Incident Solar South Wall [Wh/m2]'].sum()
      A.cell(column=init_col+3, row=row_i).value = dfh['Incident Solar West Wall [Wh/m2]'].sum()

      # Sky temperature
      A.cell(column=init_col+8, row=row_i).value = dfh['Sky Temp [C]'].mean()

    dfd = df[(df['Month'] == 7) & (df['Day'] == 14)]
    for hour in range(1,25):
      dfh = dfd[dfd["Hour"] == hour]
      row_i = 230 + hour - 1

      # Solar Incidence
      A.cell(column=init_col+4, row=row_i).value = dfh['Incident Solar Roof [Wh/m2]'].sum()
      A.cell(column=init_col+5, row=row_i).value = dfh['Incident Solar South Wall [Wh/m2]'].sum()
      A.cell(column=init_col+6, row=row_i).value = dfh['Incident Solar West Wall [Wh/m2]'].sum()

      # Sky temperature
      A.cell(column=init_col+9, row=row_i).value = dfh['Sky Temp [C]'].mean()

    dfd = df[(df['Month'] == 2) & (df['Day'] == 1)]
    for hour in range(1,25):
      dfh = dfd[dfd["Hour"] == hour]
      row_i = 230 + hour - 1

      # Sky temperature
      A.cell(column=init_col+7, row=row_i).value = dfh['Sky Temp [C]'].mean()

  # Hourly loads
  if case in col_loads:
    dfd = df[(df['Month'] == 2) & (df['Day'] == 1)]

    for hour in range(1,25):
      dfh = dfd[dfd["Hour"] == hour]
      row_i = 262 + hour - 1

      A.cell(column=init_col+col_loads[case], row=row_i).value = (dfh['Heating Load [Wh]'].sum() - dfh['Cooling Load [Wh]'].sum())/1000

    dfd = df[(df['Month'] == 7) & (df['Day'] == 14)]

    for hour in range(1,25):
      dfh = dfd[dfd["Hour"] == hour]
      row_i = 262 + hour - 1

      A.cell(column=init_col+col_loads[case]+10, row=row_i).value = (dfh['Heating Load [Wh]'].sum() - dfh['Cooling Load [Wh]'].sum())/1000

  if case in ["600", "200"]:
      # Convection
      row_i = 501

      if case == "600":
        col_offset = 0
      else:
        col_offset = 7

      ## Load weighted average
      for surf in ['South Wall Window 1','Roof','Floor','North Wall','East Wall','West Wall','South Wall']:

        # Exterior
        A.cell(column=init_col+col_offset+1, row=row_i).value = (df['Exterior Conv. Coeff. ' + surf + ' [W/m2-K]'] * (df['Heating Load [Wh]'] + df['Cooling Load [Wh]'])).sum()/(df['Heating Load [Wh]'].sum() + df['Cooling Load [Wh]'].sum())
        if case == "600":
          idx = df['Exterior Conv. Coeff. ' + surf + ' [W/m2-K]'].idxmax()
          A.cell(column=init_col+col_offset+1, row=row_i+10).value = df.ix[idx]['Exterior Conv. Coeff. ' + surf + ' [W/m2-K]']
          A.cell(column=init_col+col_offset+2, row=row_i+10).value = df.ix[idx]['Month']
          A.cell(column=init_col+col_offset+3, row=row_i+10).value = df.ix[idx]['Day']
          A.cell(column=init_col+col_offset+4, row=row_i+10).value = df.ix[idx]['Hour']

          idx = df['Exterior Conv. Coeff. ' + surf + ' [W/m2-K]'].idxmin()
          A.cell(column=init_col+col_offset+1, row=row_i+20).value = df.ix[idx]['Exterior Conv. Coeff. ' + surf + ' [W/m2-K]']
          A.cell(column=init_col+col_offset+2, row=row_i+20).value = df.ix[idx]['Month']
          A.cell(column=init_col+col_offset+3, row=row_i+20).value = df.ix[idx]['Day']
          A.cell(column=init_col+col_offset+4, row=row_i+20).value = df.ix[idx]['Hour']

        # Interior
        A.cell(column=init_col+col_offset+1, row=row_i+30).value = (df['Interior Conv. Coeff. ' + surf + ' [W/m2-K]'] * (df['Heating Load [Wh]'] + df['Cooling Load [Wh]'])).sum()/(df['Heating Load [Wh]'].sum() + df['Cooling Load [Wh]'].sum())
        if case == "600":
          idx = df['Interior Conv. Coeff. ' + surf + ' [W/m2-K]'].idxmax()
          A.cell(column=init_col+col_offset+1, row=row_i+40).value = df.ix[idx]['Interior Conv. Coeff. ' + surf + ' [W/m2-K]']
          A.cell(column=init_col+col_offset+2, row=row_i+40).value = df.ix[idx]['Month']
          A.cell(column=init_col+col_offset+3, row=row_i+40).value = df.ix[idx]['Day']
          A.cell(column=init_col+col_offset+4, row=row_i+40).value = df.ix[idx]['Hour']

          idx = df['Interior Conv. Coeff. ' + surf + ' [W/m2-K]'].idxmin()
          A.cell(column=init_col+col_offset+1, row=row_i+50).value = df.ix[idx]['Interior Conv. Coeff. ' + surf + ' [W/m2-K]']
          A.cell(column=init_col+col_offset+2, row=row_i+50).value = df.ix[idx]['Month']
          A.cell(column=init_col+col_offset+3, row=row_i+50).value = df.ix[idx]['Day']
          A.cell(column=init_col+col_offset+4, row=row_i+50).value = df.ix[idx]['Hour']

        row_i += 1

  if case == "960":
    # Sun Zone free-float
    A.cell(column=init_col+1, row=136).value = df['Sun Zone Temp [C]'].mean()
    idx = df['Sun Zone Temp [C]'].idxmin()
    A.cell(column=init_col+2, row=136).value = df.ix[idx]['Sun Zone Temp [C]']
    A.cell(column=init_col+3, row=136).value = df.ix[idx]['Month']
    A.cell(column=init_col+4, row=136).value = df.ix[idx]['Day']
    A.cell(column=init_col+5, row=136).value = df.ix[idx]['Hour']
    idx = df['Sun Zone Temp [C]'].idxmax()
    A.cell(column=init_col+6, row=136).value = df.ix[idx]['Sun Zone Temp [C]']
    A.cell(column=init_col+7, row=136).value = df.ix[idx]['Month']
    A.cell(column=init_col+8, row=136).value = df.ix[idx]['Day']
    A.cell(column=init_col+9, row=136).value = df.ix[idx]['Hour']


# free-float cases
for row in range(row_ff_beg, row_ff_end + 1):
  case = str(A.cell(column=init_col, row=row).value)
  print "  " + case
  df = pd.read_csv('../output/' + case + '/DETAILED.csv')
  A.cell(column=init_col+1, row=row).value = df['Zone Temp [C]'].mean()
  idx = df['Zone Temp [C]'].idxmin()
  A.cell(column=init_col+2, row=row).value = df.ix[idx]['Zone Temp [C]']
  A.cell(column=init_col+3, row=row).value = df.ix[idx]['Month']
  A.cell(column=init_col+4, row=row).value = df.ix[idx]['Day']
  A.cell(column=init_col+5, row=row).value = df.ix[idx]['Hour']
  idx = df['Zone Temp [C]'].idxmax()
  A.cell(column=init_col+6, row=row).value = df.ix[idx]['Zone Temp [C]']
  A.cell(column=init_col+7, row=row).value = df.ix[idx]['Month']
  A.cell(column=init_col+8, row=row).value = df.ix[idx]['Day']
  A.cell(column=init_col+9, row=row).value = df.ix[idx]['Hour']

  if case == "900FF":
    # Hourly annual temperatures
    row_tb = 11
    i = 0
    temps = np.array(df.groupby(['Month','Day','Hour'])['Zone Temp [C]'].mean())
    for month in range(1,13):
      for day in range(1,d_month[month-1]+1):
        for hour in range(1,25):
          TMPBIN.cell(column=5,row=row_tb+i).value = temps[i]
          i += 1

  if case in ["600FF","900FF","650FF","950FF","680FF","980FF"]:
    # Hourly temperatures
    if case not in ["650FF","950FF"]:
      dfd = df[(df['Month'] == 2) & (df['Day'] == 1)]
    else:
      dfd = df[(df['Month'] == 7) & (df['Day'] == 14)]

    for hour in range(1,25):
      dfh = dfd[dfd["Hour"] == hour]
      row_i = 262 + hour - 1
      A.cell(column=init_col+col_temps[case], row=row_i).value = dfh['Zone Temp [C]'].mean()

print "Done"
wb.save(filename='../reports/Sec5-2Aout.xlsx')
