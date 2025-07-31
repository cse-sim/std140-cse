print("\nInitializing Python...\n")
import openpyxl as xl
import pandas as pd
import numpy as np
import subprocess
import re
import mako.template as mk
from pathlib import Path

from typing import Dict, List

test = "space-cooling"
wb = xl.load_workbook(
    filename=Path("reports", test, "Std140_CE_a_Output-Template.xlsx")
)
A = wb["A"]

# Top level info
info = str(subprocess.check_output("CSE.exe", shell=True))
match = re.compile(".*CSE\s+([^\s]*)\s+.*", re.S).match(info)
if match:
    version = match.groups()[0]
else:
    print("ERROR: Unknown version!")
    version = "????"

FIRST_ROW = 25  # CE100
LAST_ROW = 38  # CE200

column_map: Dict[str, int] = {
    "Total_Cooling_Energy_Consumption": 2,
    "Total_Evaporator_Coil_Load": 6,
    "Total_Zone_Load": 9,
    "Mean_COP": 12,
    "Mean_Indoor_Dry_Bulb": 13,
    "Mean_Humidity_Ratio": 14,
    "Max_COP": 15,
    "Max_Indoor_Dry_Bulb": 16,
    "Max_Humidity_Ratio": 17,
    "Min_COP": 18,
    "Min_Indoor_Dry_Bulb": 19,
    "Min_Humidity_Ratio": 20,
}

cases: List[str] = [
    "CE100",
    "CE110",
    "CE120",
    "CE130",
    "CE140",
    "CE150",
    "CE160",
    "CE165",
    "CE170",
    "CE180",
    "CE185",
    "CE190",
    "CE195",
    "CE200",
]


print("Processing case: ")

# Non-free-float cases
for case, row in zip(cases, range(FIRST_ROW, LAST_ROW + 1)):
    # Cooling Energy Consumption
    print("  " + case)
    df_hourly = pd.read_csv(Path("output", test, case, "HOURLY.csv"))
    df_monthly = pd.read_csv(Path("output", test, case, "MONTHLY.csv"))

    # Annual loads
    A.cell(column=2, row=row).value = df_monthly[
        "Total Cooling Energy Consumption [kWh]"
    ].tolist()[0]
    A.cell(column=3, row=row).value = df_monthly[
        "Total Cooling Energy Consumption [kWh]"
    ].tolist()[0]
    A.cell(column=4, row=row).value = 0
    A.cell(column=5, row=row).value = 0
    A.cell(column=6, row=row).value = df_monthly[
        "Total Evaporator Coil Load [kWh]"
    ].tolist()[0]
    A.cell(column=7, row=row).value = df_monthly[
        "Sensible Evaporator Coil Load [kWh]"
    ].tolist()[0]
    A.cell(column=8, row=row).value = df_monthly[
        "Latent Evaporator Coil Load [kWh]"
    ].tolist()[0]
    A.cell(column=9, row=row).value = df_monthly[
        "Total Evaporator Coil Load [kWh]"
    ].tolist()[0]
    A.cell(column=10, row=row).value = df_monthly[
        "Sensible Evaporator Coil Load [kWh]"
    ].tolist()[0]
    A.cell(column=11, row=row).value = df_monthly[
        "Latent Evaporator Coil Load [kWh]"
    ].tolist()[0]
    A.cell(column=12, row=row).value = df_hourly["COP"].mean()
    A.cell(column=13, row=row).value = df_hourly["Zone Temp [°C]"].mean()
    A.cell(column=14, row=row).value = df_hourly["Humidity Ratio [kg/kg]"].mean()
    A.cell(column=15, row=row).value = df_hourly["COP"].max()
    A.cell(column=16, row=row).value = df_hourly["Zone Temp [°C]"].max()
    A.cell(column=17, row=row).value = df_hourly["Humidity Ratio [kg/kg]"].max()
    A.cell(column=18, row=row).value = df_hourly["COP"].min()
    A.cell(column=19, row=row).value = df_hourly["Zone Temp [°C]"].min()
    A.cell(column=20, row=row).value = df_hourly["Humidity Ratio [kg/kg]"].min()

wb.save(filename=Path("reports", test, "Std140_CE_a_Output.xlsx"))

# with open(
#     "../../reports/" + test + "/S140outNotes-Template.txt", "r"
# ) as notes_template:
#     content = notes_template.read()

# with open("../../reports/" + test + "/S140outNotes.txt", "w") as notes:
#     notes.write(mk.Template(content).render(version=version))

print("Done")
