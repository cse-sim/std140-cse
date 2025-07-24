print("\nInitializing Python...\n")

from datetime import datetime
import glob
import os
from pathlib import Path


import openpyxl as xl
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd
import pytz


def call_csv(path):
    data = pd.read_csv(path)
    return pd.DataFrame(data)


# delete output file
def delete_output_file(path):
    for file in glob.glob(f"{path}*"):
        os.remove(file)


# prints date and time with time zone on each plot
def find_todays_date():
    utc_timezone = pytz.timezone("America/Denver")
    current_date_time = datetime.now(utc_timezone)
    return current_date_time.strftime("%b %d, %Y")


template_column_row = 2
first_data_row = template_column_row + 1
last_data_row = first_data_row + 8760

m_ft = 3.28084
south_perimeter_volume = 40.7643 * 4.5732 * 2.7432

template_file_root = "Std140_CB_Output"
test_suite = "std-140"
if __name__ == "__main__":
    current_directory = os.getcwd()  # Use when running script directly
else:
    current_directory = os.path.dirname(os.path.dirname(os.getcwd()))  # Use when called from rakefile
template_file_name = f"{template_file_root}_Template_R1.xlsx"
template_file_path = Path(f"{current_directory}/docs/{test_suite}/{template_file_name}")

hourly_results_file_name = "OUTPUT_HOURLY.csv"
sub_hourly_results_file_name = "OUTPUT_SUB_HOURLY.csv"

results_directory = Path(f"{current_directory}/output/{test_suite}")

todays_date = find_todays_date()

cases = ["CB1000", "CB1010", "CB1020", "CB1100"]

annual_tab = "AnnualOutputs"
hourly_sheets = [
    "ZoneAirTemp",
    "SensibleCoolingRate",
    "HeatingRate",
    "LightingPower",
    "PlugLoadPower",
]
zone_sheet = "Hourly-Bottom_Perimeter_South"

infiltration_mass_flow_rate = "Infiltration mass flow rate [kg/s] b"
infiltration_sensible_heat_transfer_rate = "Sensible heat transfer rate into the zone due to infiltration [kW] c"
infiltration_moisture_added = "Moisture added to the zone due to the infiltration [kg water/kg dry air]"
infiltration_latent_heat_transfer_rate = "Latent heat transfer rate into the zone due to infiltration [kW] c"
ventilation_mass_flow_rate = "Ventilation mass flow rate [kg/s] b"
ventilation_sensible_heat_transfer_rate = "Sensible heat transfer rate into the zone due to ventilation [kW] c"
ventilation_moisture_added = "Moisture added to the zone due to the ventilation [kg water/kg dry air]"
ventilation_latent_heat_transfer_rate = "Latent heat transfer rate into the zone due to ventilation [kW] c"

window_net_heat_transfer_rate = "Total net heat transfer rate through the windows [kW] c,e"
# window_net_heat_transfer_rate_conduction = f"{window_net_heat_transfer_rate} Conduction"
# window_net_heat_transfer_rate_radiation = (
#     f"{window_net_heat_transfer_rate} Incident Radiation"
# )

zone_columns = [
    "Outdoor air density a [kg/m3]",
    infiltration_mass_flow_rate,
    infiltration_sensible_heat_transfer_rate,
    infiltration_moisture_added,
    infiltration_latent_heat_transfer_rate,
    ventilation_mass_flow_rate,
    ventilation_sensible_heat_transfer_rate,
    ventilation_moisture_added,
    ventilation_latent_heat_transfer_rate,
    "Total window transmitted solar radiation rate [kW] c,d",
    window_net_heat_transfer_rate,
    "Total exterior surface incident solar radiation rate [kW] d",
    "Total exterior surface convection heat transfer rate [kW] f",
    "Total interior surface convection heat transfer rate [kW] f",
    "Total interior surface convection heat transfer rate [kW] ^",
]

sub_hourly_average = [
    "Dry Air Mass",
    "ACH",
    "Moist Air Density [kg/m3]",
    "Total Sensible Heat Transfer [kW] qIzSh",
]


def post_processing(df_hourly: pd.DataFrame, df_sub_hourly: pd.DataFrame):
    df_sub_hourly["Minute"] = (df_sub_hourly["SubHour"] - 1) * 10
    df_sub_hourly["Datetime"] = pd.to_datetime(df_sub_hourly[["Month", "Day", "Hour", "Minute"]].assign(year=2024))
    df_hourly["Datetime"] = pd.to_datetime(df_hourly[["Month", "Day", "Hour"]].assign(year=2024))
    df_sub_hourly.index = df_sub_hourly["Datetime"]
    df_hourly.index = df_hourly["Datetime"]
    for column in df_sub_hourly.columns:
        if column in sub_hourly_average:
            df_hourly[column] = df_sub_hourly[column].resample("h").mean()

    df_hourly[ventilation_sensible_heat_transfer_rate] = df_hourly[ventilation_mass_flow_rate] * df_hourly["Sensible Heat Change [kJ/kg]"]
    df_hourly[ventilation_latent_heat_transfer_rate] = df_hourly[ventilation_mass_flow_rate] * df_hourly["Latent Heat Change [kJ/kg]"]

    df_hourly[infiltration_sensible_heat_transfer_rate] = df_hourly[infiltration_mass_flow_rate] * df_hourly["Sensible Heat Change [kJ/kg]"]
    df_hourly[infiltration_latent_heat_transfer_rate] = df_hourly[infiltration_mass_flow_rate] * df_hourly["Latent Heat Change [kJ/kg]"]

    df_hourly.index = [value for value in range(8760)]
    return df_hourly


for case in cases:
    template = xl.load_workbook(filename=template_file_path)

    information_sheet = template["Information"]
    information_sheet.cell(row=2, column=2, value=case)
    information_sheet.cell(row=3, column=2, value="CSE 0.925.0")
    information_sheet.cell(row=4, column=2, value="Dec 18, 2024")
    information_sheet.cell(row=5, column=2, value="Big Ladder Software")
    information_sheet.cell(
        row=6,
        column=2,
        value="neal.kruis@bigladdersoftware.com; nathan.oliver@bigladdersoftware.com; chipbarnaby@gmail.com",
    )
    information_sheet.cell(row=7, column=2, value=todays_date)

    output_file_name = f"{template_file_root}_{case}.xlsx"
    output_file_path = Path(f"{current_directory}/reports/{test_suite}/{output_file_name}")
    delete_output_file(output_file_path)

    hourly_results_file_path = Path(results_directory, case, hourly_results_file_name)
    sub_hourly_results_file_path = Path(results_directory, case, sub_hourly_results_file_name)

    df_case_data_hourly = pd.read_csv(hourly_results_file_path)
    df_case_data_sub_hourly = pd.read_csv(sub_hourly_results_file_path)

    df_case_data_hourly = post_processing(df_case_data_hourly, df_case_data_sub_hourly)

    sheets = pd.read_excel(template, sheet_name=None, engine="openpyxl", header=1)

    for hourly_sheet in hourly_sheets:
        hourly_sheet_name = f"Hourly-{hourly_sheet}"
        df_sheet = sheets[hourly_sheet_name]
        columns = df_sheet.columns[1:]
        for column in columns:
            case_data_column = f"{hourly_sheet}_{column}"
            df_sheet[column] = df_case_data_hourly[case_data_column]

        sheet = template[hourly_sheet_name]  # Access the corresponding sheet in the template
        for row_idx, row in enumerate(dataframe_to_rows(df_sheet, index=False, header=False), start=2):
            for col_idx, value in enumerate(row, start=1):
                sheet.cell(row=row_idx + 1, column=col_idx, value=value)
        print("Data successfully added to the XLSX file.")

    df_sheet = sheets[zone_sheet]
    df_sheet = df_sheet.iloc[:, :-1]
    columns = df_sheet.iloc[0]
    df_sheet = df_sheet[1:]
    df_sheet.columns = [column for column in columns]
    df_sheet.reset_index(drop=True, inplace=True)
    for zone_column in zone_columns:
        if zone_column not in [
            infiltration_moisture_added,
            ventilation_moisture_added,
        ]:  # Skips moisture-added columns since CSE calculation not performed, and not a required output.
            df_sheet[zone_column] = df_case_data_hourly[zone_column]

    sheet = template[zone_sheet]  # Access the corresponding sheet in the template
    for row_idx, row in enumerate(dataframe_to_rows(df_sheet, index=False, header=False), start=3):
        for col_idx, value in enumerate(row, start=1):
            sheet.cell(row=row_idx + 1, column=col_idx, value=value)

    print(f"Done processing case {case}.")
    print("Writing results to XLSX ...")
    template.save(f"{current_directory}/reports/{test_suite}/{output_file_name}")
print(f"\nDone processing all cases.")
