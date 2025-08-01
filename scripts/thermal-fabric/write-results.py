print("\nInitializing Python...\n")
import openpyxl as xl
import pandas as pd
import numpy as np
from datetime import datetime
import subprocess
import re
import mako.template as mk

# %matplotlib inline

tests = "thermal-fabric"
wb = xl.load_workbook(
    filename="../../reports/" + tests + "/Std140_TF_Output-Template.xlsx"
)
A = wb["A"]
TMPBIN = wb["TMPBIN"]

# Top level info
info = str(subprocess.check_output("..\\..\\CSE.exe", shell=True))
match = re.compile(".*CSE\s+([^\s]*)\s+.*", re.S).match(info)
if match:
    version = match.groups()[0]
else:
    print("ERROR: Unknown version!")
    version = "????"

A.cell(column=3, row=61).value = "CSE"
A.cell(column=3, row=62).value = version
A.cell(column=3, row=63).value = "{}".format(datetime.now())

row_beg = 70  # 600
row_end = 115  # 810
row_ff_beg = 130  # 600FF
row_ff_end = 135  # 980FF
init_col = 2  # 'B'

col_cold_loads = {
    "600": 1,
    "640": 2,
    "660": 3,
    "670": 4,
    "680": 5,
    "685": 6,
    "695": 7,
    "900": 8,
    "940": 9,
    "980": 10,
    "985": 11,
    "995": 12,
}

col_hot_loads = {
    "600": 13,
    "660": 14,
    "670": 15,
    "680": 16,
    "685": 17,
    "695": 18,
    "900": 19,
    "980": 20,
    "985": 21,
    "995": 22,
}

col_loads_max = col_hot_loads["995"]

col_cond_temps = {"640": col_loads_max + 1, "940": col_loads_max + 2}

col_temps = {"600FF": 1, "900FF": 2, "650FF": 3, "950FF": 4, "680FF": 5, "980FF": 6}

d_month = [31, 28, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31]

print("Proccessing case: ")

# Non-free-float cases
for row in range(row_beg, row_end + 1):
    case = str(A.cell(column=init_col, row=row).value)
    print("  " + case)
    df = pd.read_csv("../../output/" + tests + "/" + case + "/DETAILED.csv")

    timestep = 1.0 / (df["Subhour"].max())

    # Annual loads
    A.cell(column=init_col + 1, row=row).value = df["Heating Load [Wh]"].sum() / 1000000
    A.cell(column=init_col + 2, row=row).value = df["Cooling Load [Wh]"].sum() / 1000000
    idx = df.groupby(["Month", "Day", "Hour"]).mean()["Heating Load [Wh]"].idxmax()
    A.cell(column=init_col + 3, row=row).value = (
        df.groupby(["Month", "Day", "Hour"]).mean().loc[idx]["Heating Load [Wh]"]
        / 1000
        / timestep
    )
    A.cell(column=init_col + 4, row=row).value = idx[0]
    A.cell(column=init_col + 5, row=row).value = idx[1]
    A.cell(column=init_col + 6, row=row).value = idx[2]
    idx = df.groupby(["Month", "Day", "Hour"]).mean()["Cooling Load [Wh]"].idxmax()
    A.cell(column=init_col + 7, row=row).value = (
        df.groupby(["Month", "Day", "Hour"]).mean().loc[idx]["Cooling Load [Wh]"]
        / 1000
        / timestep
    )
    A.cell(column=init_col + 8, row=row).value = idx[0]
    A.cell(column=init_col + 9, row=row).value = idx[1]
    A.cell(column=init_col + 10, row=row).value = idx[2]

    if case == "600":
        # Solar Incidence
        A.cell(column=init_col + 1, row=155).value = (
            df["Incident Solar Roof [Wh/m2]"].sum() / 1000
        )
        A.cell(column=init_col + 1, row=156).value = (
            df["Incident Solar North Wall [Wh/m2]"].sum() / 1000
        )
        A.cell(column=init_col + 1, row=157).value = (
            df["Incident Solar East Wall [Wh/m2]"].sum() / 1000
        )
        A.cell(column=init_col + 1, row=158).value = (
            df["Incident Solar South Wall [Wh/m2]"].sum() / 1000
        )
        A.cell(column=init_col + 1, row=159).value = (
            df["Incident Solar West Wall [Wh/m2]"].sum() / 1000
        )

        # Transmitted Solar
        A.cell(column=init_col + 1, row=163).value = (
            df["Transmitted Solar [Wh/m2]"].sum() / 1000
        )

        # Sky Temperature
        A.cell(column=init_col + 1, row=178).value = df["Sky Temp [C]"].mean()
        idx = df.groupby(["Month", "Day", "Hour"]).mean()["Sky Temp [C]"].idxmin()
        A.cell(column=init_col + 2, row=178).value = (
            df.groupby(["Month", "Day", "Hour"]).mean().loc[idx]["Sky Temp [C]"]
        )
        A.cell(column=init_col + 3, row=178).value = idx[0]
        A.cell(column=init_col + 4, row=178).value = idx[1]
        A.cell(column=init_col + 5, row=178).value = idx[2]
        idx = df.groupby(["Month", "Day", "Hour"]).mean()["Sky Temp [C]"].idxmax()
        A.cell(column=init_col + 6, row=178).value = (
            df.groupby(["Month", "Day", "Hour"]).mean().loc[idx]["Sky Temp [C]"]
        )
        A.cell(column=init_col + 7, row=178).value = idx[0]
        A.cell(column=init_col + 8, row=178).value = idx[1]
        A.cell(column=init_col + 9, row=178).value = idx[2]

    if case == "660":
        # Transmitted Solar
        A.cell(column=init_col + 1, row=164).value = (
            df["Transmitted Solar [Wh/m2]"].sum() / 1000
        )

    if case == "670":
        # Transmitted Solar
        A.cell(column=init_col + 1, row=165).value = (
            df["Transmitted Solar [Wh/m2]"].sum() / 1000
        )

    if case == "620":
        # Transmitted Solar
        A.cell(
            column=init_col + 1, row=166
        ).value = "N/A"  # CSE (0.926.0) does not have a probe to measure transmitted solar through a single window. https://github.com/cse-sim/cse/issues/15

    if case == "610":
        # Transmitted Solar
        A.cell(column=init_col + 1, row=170).value = (
            df["Transmitted Solar [Wh/m2]"].sum() / 1000
        )

    if case == "630":
        # Transmitted Solar
        A.cell(
            column=init_col + 1, row=171
        ).value = "N/A"  # CSE (0.926.0) does not have a probe to measure transmitted solar through a single window. https://github.com/cse-sim/cse/issues/15

    if case == "600" or case == "900":
        # Monthly loads
        if case == "600":
            col_offset = 0
        else:
            col_offset = 8

        for month in range(1, 13):
            dfm = df[df["Month"] == month]
            A.cell(column=init_col + col_offset + 1, row=189 + month).value = (
                dfm["Heating Load [Wh]"].sum() / 1000
            )
            A.cell(column=init_col + col_offset + 2, row=189 + month).value = (
                dfm["Cooling Load [Wh]"].sum() / 1000
            )
            idx = (
                dfm.groupby(["Month", "Day", "Hour"])
                .mean()["Heating Load [Wh]"]
                .idxmax()
            )
            A.cell(column=init_col + col_offset + 3, row=189 + month).value = (
                dfm.groupby(["Month", "Day", "Hour"])
                .mean()
                .loc[idx]["Heating Load [Wh]"]
                / 1000
                / timestep
            )
            A.cell(column=init_col + col_offset + 4, row=189 + month).value = idx[1]
            A.cell(column=init_col + col_offset + 5, row=189 + month).value = idx[2]
            idx = (
                dfm.groupby(["Month", "Day", "Hour"])
                .mean()["Cooling Load [Wh]"]
                .idxmax()
            )
            A.cell(column=init_col + col_offset + 6, row=189 + month).value = (
                dfm.groupby(["Month", "Day", "Hour"])
                .mean()
                .loc[idx]["Cooling Load [Wh]"]
                / 1000
                / timestep
            )
            A.cell(column=init_col + col_offset + 7, row=189 + month).value = idx[1]
            A.cell(column=init_col + col_offset + 8, row=189 + month).value = idx[2]

    # Hourly outputs (misc.)
    if case == "600":
        dfd = df[(df["Month"] == 5) & (df["Day"] == 4)]
        for hour in range(1, 25):
            dfh = dfd[dfd["Hour"] == hour]
            row_i = 230 + hour - 1

            # Solar Incidence
            A.cell(column=init_col + 1, row=row_i).value = dfh[
                "Incident Solar Roof [Wh/m2]"
            ].sum()
            A.cell(column=init_col + 2, row=row_i).value = dfh[
                "Incident Solar South Wall [Wh/m2]"
            ].sum()
            A.cell(column=init_col + 3, row=row_i).value = dfh[
                "Incident Solar West Wall [Wh/m2]"
            ].sum()

            # Sky temperature
            A.cell(column=init_col + 8, row=row_i).value = dfh["Sky Temp [C]"].mean()

            # Transmitted Solar
            A.cell(column=init_col + 13, row=row_i).value = dfh[
                "Transmitted Solar [Wh/m2]"
            ].sum()

        dfd = df[(df["Month"] == 7) & (df["Day"] == 14)]
        for hour in range(1, 25):
            dfh = dfd[dfd["Hour"] == hour]
            row_i = 230 + hour - 1

            # Solar Incidence
            A.cell(column=init_col + 4, row=row_i).value = dfh[
                "Incident Solar Roof [Wh/m2]"
            ].sum()
            A.cell(column=init_col + 5, row=row_i).value = dfh[
                "Incident Solar South Wall [Wh/m2]"
            ].sum()
            A.cell(column=init_col + 6, row=row_i).value = dfh[
                "Incident Solar West Wall [Wh/m2]"
            ].sum()

            # Sky temperature
            A.cell(column=init_col + 9, row=row_i).value = dfh["Sky Temp [C]"].mean()

            # Transmitted Solar
            A.cell(column=init_col + 16, row=row_i).value = dfh[
                "Transmitted Solar [Wh/m2]"
            ].sum()

        dfd = df[(df["Month"] == 2) & (df["Day"] == 1)]
        for hour in range(1, 25):
            dfh = dfd[dfd["Hour"] == hour]
            row_i = 230 + hour - 1

            # Sky temperature
            A.cell(column=init_col + 7, row=row_i).value = dfh["Sky Temp [C]"].mean()

            # Transmitted Solar
            A.cell(column=init_col + 10, row=row_i).value = dfh[
                "Transmitted Solar [Wh/m2]"
            ].sum()

    if case == "660":
        dfd = df[(df["Month"] == 5) & (df["Day"] == 4)]
        for hour in range(1, 25):
            dfh = dfd[dfd["Hour"] == hour]
            row_i = 230 + hour - 1

            # Transmitted Solar
            A.cell(column=init_col + 14, row=row_i).value = dfh[
                "Transmitted Solar [Wh/m2]"
            ].sum()

        dfd = df[(df["Month"] == 7) & (df["Day"] == 14)]
        for hour in range(1, 25):
            dfh = dfd[dfd["Hour"] == hour]
            row_i = 230 + hour - 1

            # Transmitted Solar
            A.cell(column=init_col + 17, row=row_i).value = dfh[
                "Transmitted Solar [Wh/m2]"
            ].sum()

        dfd = df[(df["Month"] == 2) & (df["Day"] == 1)]
        for hour in range(1, 25):
            dfh = dfd[dfd["Hour"] == hour]
            row_i = 230 + hour - 1

            # Transmitted Solar
            A.cell(column=init_col + 11, row=row_i).value = dfh[
                "Transmitted Solar [Wh/m2]"
            ].sum()

    if case == "670":
        dfd = df[(df["Month"] == 5) & (df["Day"] == 4)]
        for hour in range(1, 25):
            dfh = dfd[dfd["Hour"] == hour]
            row_i = 230 + hour - 1

            # Transmitted Solar
            A.cell(column=init_col + 15, row=row_i).value = dfh[
                "Transmitted Solar [Wh/m2]"
            ].sum()

        dfd = df[(df["Month"] == 7) & (df["Day"] == 14)]
        for hour in range(1, 25):
            dfh = dfd[dfd["Hour"] == hour]
            row_i = 230 + hour - 1

            # Transmitted Solar
            A.cell(column=init_col + 18, row=row_i).value = dfh[
                "Transmitted Solar [Wh/m2]"
            ].sum()

        dfd = df[(df["Month"] == 2) & (df["Day"] == 1)]
        for hour in range(1, 25):
            dfh = dfd[dfd["Hour"] == hour]
            row_i = 230 + hour - 1

            # Transmitted Solar
            A.cell(column=init_col + 12, row=row_i).value = dfh[
                "Transmitted Solar [Wh/m2]"
            ].sum()

    # Hourly loads
    if case in col_cold_loads:
        dfd = df[(df["Month"] == 2) & (df["Day"] == 1)]

        for hour in range(1, 25):
            dfh = dfd[dfd["Hour"] == hour]
            row_i = 262 + hour - 1

            A.cell(column=init_col + col_cold_loads[case], row=row_i).value = (
                dfh["Heating Load [Wh]"].sum() - dfh["Cooling Load [Wh]"].sum()
            ) / 1000

    if case in col_hot_loads:
        dfd = df[(df["Month"] == 7) & (df["Day"] == 14)]

        for hour in range(1, 25):
            dfh = dfd[dfd["Hour"] == hour]
            row_i = 262 + hour - 1

            A.cell(column=init_col + col_hot_loads[case], row=row_i).value = (
                dfh["Heating Load [Wh]"].sum() - dfh["Cooling Load [Wh]"].sum()
            ) / 1000

    if case in col_cond_temps:
        dfd = df[(df["Month"] == 2) & (df["Day"] == 1)]

        for hour in range(1, 25):
            dfh = dfd[dfd["Hour"] == hour]
            row_i = 262 + hour - 1
            A.cell(column=init_col + col_cond_temps[case], row=row_i).value = dfh[
                "Zone Temp [C]"
            ].mean()

    if case in ["600", "200", "670"]:
        # Convection
        row_i = 501

        if case == "600":
            col_offset = 0
        elif case == "200":
            col_offset = 7
        else:
            col_offset = 10

        ## Load weighted average
        for surf in [
            "South Wall Window 1",
            "Roof",
            "Floor",
            "North Wall",
            "East Wall",
            "West Wall",
            "South Wall",
        ]:
            if case == "670" and surf == "Roof":
                break

            # Exterior
            A.cell(column=init_col + col_offset + 1, row=row_i).value = (
                df["Exterior Conv. Coeff. " + surf + " [W/m2-K]"]
                * (df["Heating Load [Wh]"] + df["Cooling Load [Wh]"])
            ).sum() / (df["Heating Load [Wh]"].sum() + df["Cooling Load [Wh]"].sum())
            if case == "600":
                idx = (
                    df.groupby(["Month", "Day", "Hour"])
                    .mean()["Exterior Conv. Coeff. " + surf + " [W/m2-K]"]
                    .idxmax()
                )
                A.cell(column=init_col + col_offset + 1, row=row_i + 10).value = (
                    df.groupby(["Month", "Day", "Hour"])
                    .mean()
                    .loc[idx]["Exterior Conv. Coeff. " + surf + " [W/m2-K]"]
                )
                A.cell(column=init_col + col_offset + 2, row=row_i + 10).value = idx[0]
                A.cell(column=init_col + col_offset + 3, row=row_i + 10).value = idx[1]
                A.cell(column=init_col + col_offset + 4, row=row_i + 10).value = idx[2]

                idx = (
                    df.groupby(["Month", "Day", "Hour"])
                    .mean()["Exterior Conv. Coeff. " + surf + " [W/m2-K]"]
                    .idxmin()
                )
                A.cell(column=init_col + col_offset + 1, row=row_i + 20).value = (
                    df.groupby(["Month", "Day", "Hour"])
                    .mean()
                    .loc[idx]["Exterior Conv. Coeff. " + surf + " [W/m2-K]"]
                )
                A.cell(column=init_col + col_offset + 2, row=row_i + 20).value = idx[0]
                A.cell(column=init_col + col_offset + 3, row=row_i + 20).value = idx[1]
                A.cell(column=init_col + col_offset + 4, row=row_i + 20).value = idx[2]

            # Interior
            A.cell(column=init_col + col_offset + 1, row=row_i + 30).value = (
                df["Interior Conv. Coeff. " + surf + " [W/m2-K]"]
                * (df["Heating Load [Wh]"] + df["Cooling Load [Wh]"])
            ).sum() / (df["Heating Load [Wh]"].sum() + df["Cooling Load [Wh]"].sum())
            if case == "600":
                idx = (
                    df.groupby(["Month", "Day", "Hour"])
                    .mean()["Interior Conv. Coeff. " + surf + " [W/m2-K]"]
                    .idxmax()
                )
                A.cell(column=init_col + col_offset + 1, row=row_i + 40).value = (
                    df.groupby(["Month", "Day", "Hour"])
                    .mean()
                    .loc[idx]["Interior Conv. Coeff. " + surf + " [W/m2-K]"]
                )
                A.cell(column=init_col + col_offset + 2, row=row_i + 40).value = idx[0]
                A.cell(column=init_col + col_offset + 3, row=row_i + 40).value = idx[1]
                A.cell(column=init_col + col_offset + 4, row=row_i + 40).value = idx[2]

                idx = (
                    df.groupby(["Month", "Day", "Hour"])
                    .mean()["Interior Conv. Coeff. " + surf + " [W/m2-K]"]
                    .idxmin()
                )
                A.cell(column=init_col + col_offset + 1, row=row_i + 50).value = (
                    df.groupby(["Month", "Day", "Hour"])
                    .mean()
                    .loc[idx]["Interior Conv. Coeff. " + surf + " [W/m2-K]"]
                )
                A.cell(column=init_col + col_offset + 2, row=row_i + 50).value = idx[0]
                A.cell(column=init_col + col_offset + 3, row=row_i + 50).value = idx[1]
                A.cell(column=init_col + col_offset + 4, row=row_i + 50).value = idx[2]

            row_i += 1

    if case == "960":
        # Sun Zone free-float
        A.cell(column=init_col + 1, row=136).value = df["Sun Zone Temp [C]"].mean()
        idx = df.groupby(["Month", "Day", "Hour"]).mean()["Sun Zone Temp [C]"].idxmin()
        A.cell(column=init_col + 2, row=136).value = (
            df.groupby(["Month", "Day", "Hour"]).mean().loc[idx]["Sun Zone Temp [C]"]
        )
        A.cell(column=init_col + 3, row=136).value = idx[0]
        A.cell(column=init_col + 4, row=136).value = idx[1]
        A.cell(column=init_col + 5, row=136).value = idx[2]
        idx = df.groupby(["Month", "Day", "Hour"]).mean()["Sun Zone Temp [C]"].idxmax()
        A.cell(column=init_col + 6, row=136).value = (
            df.groupby(["Month", "Day", "Hour"]).mean().loc[idx]["Sun Zone Temp [C]"]
        )
        A.cell(column=init_col + 7, row=136).value = idx[0]
        A.cell(column=init_col + 8, row=136).value = idx[1]
        A.cell(column=init_col + 9, row=136).value = idx[2]


# free-float cases
for row in range(row_ff_beg, row_ff_end + 1):
    case = str(A.cell(column=init_col, row=row).value)
    print("  " + case)
    df = pd.read_csv("../../output/" + tests + "/" + case + "/DETAILED.csv")
    A.cell(column=init_col + 1, row=row).value = df["Zone Temp [C]"].mean()
    idx = df.groupby(["Month", "Day", "Hour"]).mean()["Zone Temp [C]"].idxmin()
    A.cell(column=init_col + 2, row=row).value = (
        df.groupby(["Month", "Day", "Hour"]).mean().loc[idx]["Zone Temp [C]"]
    )
    A.cell(column=init_col + 3, row=row).value = idx[0]
    A.cell(column=init_col + 4, row=row).value = idx[1]
    A.cell(column=init_col + 5, row=row).value = idx[2]
    idx = df.groupby(["Month", "Day", "Hour"]).mean()["Zone Temp [C]"].idxmax()
    A.cell(column=init_col + 6, row=row).value = (
        df.groupby(["Month", "Day", "Hour"]).mean().loc[idx]["Zone Temp [C]"]
    )
    A.cell(column=init_col + 7, row=row).value = idx[0]
    A.cell(column=init_col + 8, row=row).value = idx[1]
    A.cell(column=init_col + 9, row=row).value = idx[2]

    if case == "900FF":
        # Hourly annual temperatures
        row_tb = 11
        i = 0
        temps = np.array(df.groupby(["Month", "Day", "Hour"])["Zone Temp [C]"].mean())
        for month in range(1, 13):
            for day in range(1, d_month[month - 1] + 1):
                for hour in range(1, 25):
                    TMPBIN.cell(column=5, row=row_tb + i).value = temps[i]
                    i += 1

    if case in ["600FF", "900FF", "650FF", "950FF", "680FF", "980FF"]:
        # Hourly temperatures
        if case not in ["650FF", "950FF"]:
            dfd = df[(df["Month"] == 2) & (df["Day"] == 1)]
        else:
            dfd = df[(df["Month"] == 7) & (df["Day"] == 14)]

        for hour in range(1, 25):
            dfh = dfd[dfd["Hour"] == hour]
            row_i = 294 + hour - 1
            A.cell(column=init_col + col_temps[case], row=row_i).value = dfh[
                "Zone Temp [C]"
            ].mean()

wb.save(filename="../../reports/" + tests + "/Std140_TF_Output.xlsx")

with open(
    "../../reports/" + tests + "/S140outNotes-Template.txt", "r"
) as notes_template:
    content = notes_template.read()

with open("../../reports/" + tests + "/S140outNotes.txt", "w") as notes:
    notes.write(mk.Template(content).render(version=version))

print("Done")
